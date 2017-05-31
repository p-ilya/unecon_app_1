import openpyxl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException


class UneconParser():
    '''
    Считывает информацию о парах
    для всех групп
    с всех рабочих листов
    одного файла.
    load_file - загрузить файл
    find_corner_cell - найти опорный угол шапки таблицы на листе
    find_start_parse - найти опорный угол тела таблицы
    set_coordinates - найти и сохранить опорные координаты
    for_group - выдать расписание для одной группы.
    '''
    keywords = ('день недели', 'пара', 'часы')
    WEEKDAYS = {'понедельник': 0,
                'вторник': 1,
                'среда': 2,
                'четверг': 3,
                'пятница': 4,
                'суббота': 5,
                'воскресенье': 6}

    current_c_num = 0

    test_name_list = []

    def load_file(self, path):
        self.wb = openpyxl.load_workbook(
            path,
            read_only=False,
            guess_types=True)
        self.worksheets = self.wb.get_sheet_names()

    def find_corner_cell(self, ws):
        #  finding left upper table corner
        min_row = 1
        max_row = 20
        max_col = 3
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
            for cell in row:
                if str(cell.value).lower() == self.keywords[0]:
                    col1 = cell.column
                    row1 = cell.row
                    self.corner_coords = (row1, column_index_from_string(col1))
                    break
                else:
                    continue
            try:
                row1
                return
            except NameError:
                continue

        # self.corner_coords = (row1, column_index_from_string(col1))

    def find_parse_start(self, ws):
        #  finding where the actual data is stored
        min_row = 1
        max_row = 20
        max_col = 3
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
            for cell in row:
                if str(cell.value).lower() == 'понедельник':
                    col1 = cell.column
                    row1 = cell.row
                    self.start_coords = (row1, column_index_from_string(col1))
                    break
                else:
                    continue
            try:
                col1
                return
            except NameError:
                continue

    def set_coordinates(self, ws):
        self.find_corner_cell(ws=ws)
        self.find_parse_start(ws=ws)

        self.WEEKDAY_COL = self.corner_coords[1]
        self.CLASSNUM_COL = self.corner_coords[1] + 1
        self.CLASSTIME_col = self.corner_coords[1] + 2
        self.GROUPNAME_ROW = self.start_coords[0] - 1
        self.PARSE_ROW = self.start_coords[0]
        self.PARSE_COL = self.start_coords[1] + 4  # main blocks +=2

        # print(self.corner_coords)
        # print(self.start_coords)

    def for_group(self, group_col, PARSE_ROW, ws):
        group_name = ws.cell(row=PARSE_ROW - 1, column=group_col).value

        if group_name is None:
            return
        if group_name.lower() in self.keywords:
            # print('returned')
            return
        if group_name == 'Пара':
            # print('raised')
            raise StopIteration
        # print('\nГруппа: '+str(group_name))
        #  print('РАСПИСАНИЕ для группы {}\n'.format(group_name))
        self.current_c_num = 0
        for row in ws.iter_rows(min_row=PARSE_ROW,
                                max_col=group_col,
                                max_row=400):
            day = row[0].value
            if not day:
                #  Конец колонки
                return False
            if day.lower() in self.WEEKDAYS.keys():
                weekday = self.WEEKDAYS[day.lower()]
            name = row[group_col - 1].value
            c_num = row[self.CLASSNUM_COL - 1].value
            if not name:
                self.current_c_num = c_num
                continue
            place = row[group_col - 2].value
            if c_num == self.current_c_num:
                week = 0  # четная неделя
            else:
                week = 1  # нечетная
                self.current_c_num = c_num
            time = row[2].value

            yield (weekday, week, time, place, name, group_name)


#  пример использования
def get_parse_generators(filename):
    '''
    Функция для всего файла.
    Проходит по всем листам и всем
    группам, для каждой получает
    генератор кортежей с данными.

    :param filename: имя файла xlsx
    :return: список генераторов
    '''
    geners = []
    p = UneconParser()
    try:
        p.load_file(filename)
        print('Файл {} загружен в парсер.'.format(filename))
    except InvalidFileException as e:
        print('Файл {} формата xls или поврежден.\nОшибка: {}'.format(filename, e))
        return None
    except OSError as e:
        print('Файл {} не найден.\nОшибка: {}'.format(filename, e))
        return None
    for ws_name in p.worksheets:
        ws = p.wb[ws_name]
        p.set_coordinates(ws=ws)

        for group_col in range(p.PARSE_COL, 400, 2):
            try:
                geners.append(p.for_group(
                    group_col=group_col,
                    PARSE_ROW=p.PARSE_ROW,
                    ws=ws))

            except AttributeError:
                break
    return geners


def parse_xlsx(filename):
    '''
    Функция получает генераторы значений
    для всего файла и формирует
    итоговый датасет в виде списка кортежей.

    :param filename: имя файла xlsx
    :return: dataset[(),(),...]
    '''
    f = get_parse_generators(filename)
    if f is None:
        return None
    dataset = []
    count = 0
    for g in f:
        for i in g:
            dataset.append(i)
            count += 1
    print('В файле {} найдено {} учебных занятий.'.format(filename, str(count)))
    return dataset


if __name__ == '__main__':
    print(parse_xlsx('files/f_ekonomiki_i_finansov_ofo_4_kurs_s_12.01.2017_g_NEW.xlsx'))
#  TO DO  **********************************************
#  xls support

