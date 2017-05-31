from zipfile import BadZipfile
import openpyxl
import xlrd, xlwt
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string as col_ifs
from openpyxl.utils.exceptions import InsufficientCoordinatesException

#filename = 'fupravleniya_ofo_1_k'
#wb1 = load_workbook('{}.xlsx'.format(filename), read_only=False, guess_types=True)


def find_parse_start(ws):
    #  finding where the actual data is stored
    find_range = ws['A1:C20']
    for row in find_range:
        for cell in row:
            if str(cell.value).lower() == 'понедельник':
                col1 = cell.column
                row1 = cell.row
                break
            else:
                continue
        try:
            row1
            break
        except NameError:
            continue

    return ('{0}{1}'.format(col1, row1))


def unmerge_xls_worksheet(ws):
    for crange in ws.merged_cells:
        rlo, rhi, clo, chi = crange
        # copying data
        col_num = chi - clo
        if col_num > 1 and clo != 0:
            data = ws.cell_value(clo, rlo)
            place = ws.cell_value(clo - 1, rlo)
            for c in range(clo, chi + 1, 2):
                for r in range(rlo, rhi + 1):
                    ws.cell(c, r).value = data
            for c in range(clo + 1, chi + 1, 2):
                for r in range(rlo, rhi + 1):
                    ws.cell(c, r).value = place
        else:
            for c in range(clo, chi + 1):
                for r in range(rlo, rhi + 1):
                    ws.cell(c, r).value = data


def unmerge_worksheet(ws):
    # corner = find_parse_start(ws)

    for diapason in ws.merged_cell_ranges:
        first, last = diapason.split(':')
        data = ws[first].value
        try:
            ws.unmerge_cells(diapason)
            # print('Unmerged: {0}'.format(diapason))

            #  copying the data
            col_num = col_ifs(ws[last].column) - col_ifs(ws[first].column)
            if col_num > 1 and 'A' not in first:
                place = ws.cell(
                    row=ws[first].row,
                    column=col_ifs(ws[first].column) - 1).value
                for i in range(0, col_num + 1, 2):
                    for cell_range in ws[first:last]:
                        cell_range[i].value = data
                for i in range(1, col_num + 1, 2):
                    for cell_range in ws[first:last]:
                        cell_range[i].value = place

            else:
                for cell_range in ws[first:last]:
                    for cell in cell_range:
                        cell.value = data

        except InsufficientCoordinatesException:
            print('These cells are not merged: {}'.format(diapason))


def workbook_unmerge(fname):
    try:
        wb1 = load_workbook(fname,
                            read_only=False,
                            keep_vba=False,
                            data_only=True,
                            guess_types=True,
                            keep_links=False)
    except OSError:
        print('Файл {} не найден.'.format(fname))
        return None
    except BadZipfile:
        print('Файл {} повреждён.'.format(fname))
        return None

    for sheet in wb1:
        unmerge_worksheet(sheet)
    new_fname = fname.split('.xlsx')[0] + '_NEW.xlsx'
    wb1.save(new_fname)
    print('Файл {} успешно преобразован.'.format(fname))
    return new_fname

'''
def xls_workbook_unmerge(fname):
    wb_xls = xlrd.open_workbook(fname+'.xls', formatting_info=True)
    sh_xls = wb_xls.sheets()
    new = xlwt.Workbook()
    
    for ws in sh_xls:
        unmerge_xls_worksheet(ws)
'''

#xls_workbook_unmerge(filename)
